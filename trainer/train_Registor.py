#!/usr/bin/python3

import argparse
import itertools
import os
import torchvision.transforms as transforms
from torch.utils.data import DataLoader
from PIL import Image
import torch
from .utils import LambdaLR,Logger,ReplayBuffer,ToTensor,Resize3D
import torch.nn.functional as F
from .utils import Logger
import numpy as np
from .Reg_datasets import ImageDataset,TestDataset,ImageDataset_brats,TestDataset_brats
from model.Eva_model import Evaluator
from model.Reg_model import VxmDense
import SimpleITK as sitk
import glob
from .utils import Transformer_3D,smooth_loss,HD, Get_Ja, NJ_loss,ASSD
import openpyxl


class Reg_Trainer():
    def __init__(self, config):
        super().__init__()
        self.config = config
        ## def networks
        self.net_R = VxmDense(cfg = config).cuda()
        # Continue training
        #self.net_R.load_state_dict(torch.load(self.config['save_root'] + 'Registration.pth'))
        self.trans_image = Transformer_3D(mode='bilinear').cuda()
        self.trans_label = Transformer_3D(mode = 'nearest').cuda()
        self.optimizer_R = torch.optim.Adam(self.net_R.parameters(), lr=config['lr'], betas=(0.5, 0.999))
        self.net_E = Evaluator(config['input_nc'], config['output_nc']).cuda()
        self.net_E.load_state_dict(torch.load(self.config['evaluator_root']))
        for (name, param) in self.net_E.named_parameters():
            param.requires_grad = False
   
        self.transforms_1 = [ ToTensor(),
                       Resize3D(size_tuple = config['size'], mode="trilinear")
                       ]
        self.transforms_mask = [ToTensor(),
                             Resize3D(size_tuple=config['size'], mode="nearest")
                             ]
        
        self.dataloader = DataLoader(ImageDataset(config['dataroot'],transforms_=self.transforms_1,
                           opt = config, unaligned=False),batch_size=config['batchSize'],shuffle=True,num_workers=config['n_cpu'])
        self.test_dataloader = DataLoader(TestDataset(config['testroot'],transforms_=self.transforms_1,
                           opt = config, unaligned=False),batch_size=1,shuffle=False,num_workers=config['n_cpu'])
        
        self.logger = Logger(config['name'],config['port'],config['n_epochs'], len(self.dataloader), config['epoch'])
        self.simlog = []
        self.smoothlog = []

    def train(self):
        simlog = self.simlog
        smoothlog = self.smoothlog
        ###### Training ######
        for epoch in range(self.config['epoch'], self.config['n_epochs']+1):
            simlog = []
            smoothlog = []
            for A,B in self.dataloader:
                # use B replace the A1_noise 
                A = A.cuda()
                B = B.cuda()
                Depth = A.shape[2]
                self.optimizer_R.zero_grad()
               
                # A regist to B
                flow = self.net_R(A, B)
                A_warp = self.trans_image(A, flow)
                error_map = self.net_E(torch.cat([B, A_warp], 1))
                error_map = torch.abs(error_map)

                # loss
                loss_sim = self.config["sim_w"] * torch.mean(error_map)
                loss_smooth = self.config["smooth_w"] * smooth_loss(flow)
                loss_reg = loss_sim + loss_smooth

                loss_reg.backward()
                self.optimizer_R.step()
                simlog.append(loss_sim)
                smoothlog.append(loss_smooth)

                self.logger.log({'L_Sim': sum(simlog)/len(simlog), 'L_Smooth': sum(smoothlog)/len(smoothlog)},
                           images={'A': A[0,:,int(Depth/2),:,:], 'B': B[0,:,int(Depth/2),:,:],                                                               'A_warp': A_warp[0,:,int(Depth/2),:,:],#'B_warp': B_warp[0,:,int(Depth/2),:,:],
                                   'error_map': error_map[0,:,int(Depth/2),:,:]})

            torch.save(self.net_R.state_dict(), self.config['save_root'] + 'Registration.pth')


    def compute_label_dice(self, A, B):
        A = A.round()
        B = B.round()
        # 需要计算的标签类别，不包括背景和图像中不存在的区域
        cls_lst = np.unique(B)[1:]
        dice_lst = []
        for cls in cls_lst:
            dice = self.cal_dice_label(A == cls, B == cls)
            dice_lst.append(dice)
            #print(dice_lst)
        return np.mean(dice_lst), dice_lst

    def test_regisatration(self):
        root_dir = self.config['save_root']
        modelname = os.path.join(root_dir, 'Registration.pth')

        Trans = transforms.Compose(self.transforms_1)
        Trans_mask = transforms.Compose(self.transforms_mask)

        # log
        log_name = "result_test_ct"
        root_path = "/opt/data/private/dataset/clinic_mrct/result_test/"
        print("log_name: ", root_path)
        f = open(os.path.join(root_path, log_name + ".txt"), "w")
        wb1 = openpyxl.Workbook()
        wb2 = openpyxl.Workbook()
        sheet_dice = wb1.active
        sheet_hd = wb2.active

        self.net_R.load_state_dict(torch.load('/opt/data/private/SynMSE-main/output/Reg/Registration.pth'))  # Replace it with the path to your registration model.
        self.net_R.eval()
        #the dice and hd95 befor deformation registration
        Befor_DICE, Befor_HD, Befor_DICE_lst, Befor_HD_lst = 0, 0, [], []
        #the result of after registration
        A2B_DICE = 0
        A2B_HD95 = 0
        num = 0
        CT_files_root = sorted(glob.glob(os.path.join(self.config['testroot'], '*', '*', 'CT_img' + '*.nii.gz')))
        MR_files_root = sorted(glob.glob(os.path.join(self.config['testroot'], '*', '*', 'MR_img' + '*.nii.gz')))
        CT_masks_root = sorted(glob.glob(os.path.join(self.config['testroot'], '*', '*', 'CT_label' + '*.nii.gz')))
        MR_masks_root = sorted(glob.glob(os.path.join(self.config['testroot'], '*', '*', 'MR_label' + '*.nii.gz')))
        print(len(MR_files_root))
        for j in range(len(MR_files_root)):
            MR_img_root = MR_files_root[j]
            MR_msk_root = MR_masks_root[j]
            CT_img_root = CT_files_root[j]
            CT_msk_root = CT_masks_root[j]

            MR_img_ori = sitk.ReadImage(MR_img_root)
            MR_msk_ori = sitk.ReadImage(MR_msk_root)
            CT_img_ori = sitk.ReadImage(CT_img_root)
            CT_msk_ori = sitk.ReadImage(CT_msk_root)

            MR_arr = sitk.GetArrayFromImage(MR_img_ori).astype(np.float32)
            MR_msk = sitk.GetArrayFromImage(MR_msk_ori).astype(np.float32)
            CT_arr = sitk.GetArrayFromImage(CT_img_ori).astype(np.float32)
            CT_msk = sitk.GetArrayFromImage(CT_msk_ori).astype(np.float32)

            CT_arr = np.clip(CT_arr, -180, 400)
            b = np.percentile(MR_arr, 99.7)
            t = np.percentile(MR_arr, 0.5)
            MR_arr = np.clip(MR_arr, t, b)

            CT_arr = 2 * (CT_arr - np.min(CT_arr)) / (np.max(CT_arr) - np.min(CT_arr)) - 1
            MR_arr = 2 * (MR_arr - np.min(MR_arr)) / (np.max(MR_arr) - np.min(MR_arr)) - 1

            CT_arr = Trans(CT_arr)
            MR_arr = Trans(MR_arr)
            CT_msk = Trans_mask(CT_msk)
            MR_msk = Trans_mask(MR_msk)


            RA = MR_arr
            RB = CT_arr
            A_mask = MR_msk
            B_mask = CT_msk

            '''
            RA = CT_arr
            RB = MR_arr
            A_mask = CT_msk
            B_mask = MR_msk

            '''
            RA = RA.cuda()
            RB = RB.cuda()
            A_mask = A_mask.cuda()
            B_mask = B_mask.cuda()
            A_mask1 = A_mask.squeeze().detach().cpu().numpy()
            B_mask1 = B_mask.squeeze().detach().cpu().numpy()
            A_mask1 = sitk.GetArrayFromImage(A_mask)
            B_mask1 = sitk.GetArrayFromImage(B_mask)

            befor_dice, befor_dice_lst = self.compute_label_dice(A_mask1, B_mask1)
            befor_hd, befor_hd_lst = HD(A_mask1, B_mask1)

            flow = self.net_R(RA, RB)

            warp2B_mask = self.trans_label(A_mask, flow)
            warped_2B = self.trans_image(RA, flow)
            a2bjdet = NJ_loss(flow)

            warp2B_mask1 = warp2B_mask.squeeze().detach().cpu().numpy().astype(np.float32)

            a2b_dice, a2b_dice_lst = self.compute_label_dice(warp2B_mask1, B_mask1)
            a2b_HD95, a2b_HD95_lst = HD(warp2B_mask1, B_mask1)

            Befor_DICE += befor_dice
            Befor_HD += befor_hd
            Befor_DICE_lst += befor_dice_lst
            Befor_HD_lst += befor_hd_lst
            A2B_DICE += a2b_dice
            A2B_HD95 += a2b_HD95

            num += 1

            print("number:", num, "befor_dice:", befor_dice, "befor_dice_lst:", befor_dice_lst)
            print("a2b_dice:", a2b_dice, "a2b_dice_lst:", a2b_dice_lst)
            print("number:", num, "befor_hd:", befor_hd, "befor_hd_lst:", befor_hd_lst)
            print("a2b_hd:", a2b_HD95, "a2b_hd_lst:", a2b_HD95_lst)
            for i in range(len(a2b_dice_lst)):
                sheet_dice.cell(row=i + 1, column=num + 1, value=a2b_dice_lst[i])
            for i in range(len(a2b_HD95_lst)):
                sheet_hd.cell(row=i + 1, column=num + 1, value=a2b_HD95_lst[i])


        print ('Befor DC:',Befor_DICE/num)
        print ('Befor HD:',Befor_HD/num)
        print ('A2B DC:',A2B_DICE/num)
        print ('A2B HD95:',A2B_HD95/num)
        print("%f, %f" % (Befor_DICE / num, Befor_HD / num), file=f)
        print("%f, %f" % (A2B_DICE/num, A2B_HD95/num), file=f)
        wb1.save(root_path + "result_dice.xlsx")
        wb2.save(root_path + "result_hd.xlsx")
                
    
###################################
if __name__ == '__main__':
    main()
